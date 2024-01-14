from datetime import date
from datetime import datetime

import openpyxl
import requests
from bs4 import BeautifulSoup

import logging_code
import Calendar
from Config import host, user, port, password, db_name, name_groups, name_teacher
import SQL_Connected

class Parsing:
    def __init__(self, program):
        #self.url_site = 'https://rksi.ru/schedule'
        self.url_site = 'https://www.rksi.ru/mobile_schedule'
        self.__log_bot = logging_code.tele_bot()
        self.calendar = Calendar.Google_calendar(program)

        self.name_groups = name_groups
        self.name_teacher = name_teacher
        self.events = []

        # try:
        #     self.__db = SQL_Connected.SQL_Connection_DB(host=host, user=user, port=port, password=password, db_name=db_name)
        #     self.__db.connection()
        # except Exception as e:
        #     self.__log_bot.message_bot(e)

    def par_group(self):
        # self.__db.connection()
        self.__log_bot.message_bot('Парсинг расписания групп  и занесения в календарь начал свою работу')
        res = requests.get(self.url_site)
        soup = BeautifulSoup(res.text, 'lxml')
        lecturer = soup.find('select', id='group').find_all('option')
        clear_list = []
        for i in lecturer:
            clear_list.append(i.text)  # Засовываем все группы в список
        print(clear_list)

        preps = []
        for group in self.name_groups:
            if group in clear_list:
                trash_response_text = requests.post(self.url_site, {'group': f'{group}'.encode('cp1251'), 'stt': 'Показать!'.encode('cp1251')}).text
                trash_prep_soup = BeautifulSoup(trash_response_text, 'lxml')

                group_list = []
                for i in trash_prep_soup.find_all(['p', 'b']):
                    group_list.append(
                        str(i).replace('<br/><b>', '||').replace('</b><br/>', '||').replace('<p>', '').replace('<b>', '').replace('</p>', '').replace('</b>', ''))

                preps.append(f'&{group}')
                for i in group_list:
                    if '</' not in i:
                        preps.append(i.split('||'))

        today = str(date.today())  # Текущая дата
        days = list()  # Засовываем дату (год-месяц-день) текущей пары
        groups = str()  # Для хранения группы
        for name in preps:  # Обрабатываем данные
            if name[0] == "&":
                groups = f"{name[1:]}"
            if 12 <= len(name[0]) <= 23 and name[0][0:1].isdigit():
                nd = name[0].split(" ")  # Разделяем дату на 'день', 'месяц', 'день недели'
                nd[1] = nd[1].replace(",", "")  # Избавляемся от знака , в месяце и засовываем в список
                month = {'января': '01', 'февраля': '02', 'марта': '03', 'апреля': '04', 'мая': '05', 'июня': '06',
                         'июля': '07', 'августа': '08', 'сентября': '09', 'октября': '10', 'ноября': '11',
                         'декабря': '12'}

                if len(nd[0]) == 1 and 1 <= int(nd[0]) < 10:
                    days.append(f"{today[0:4]}-{month[nd[1]]}-0{nd[0]}")
                elif len(nd[0]) == 2 and int(nd[0]) >= 10:
                    days.append(f"{today[0:4]}-{month[nd[1]]}-{nd[0]}")

            try:
                if len(name) == 3:  # Обработанные данные загружаем в календарь
                    print(f'{name[1]}', f'Время: {name[0]}: {name[2]}', f'{days[-1]}', f'{name[0]}', '7', groups)
                    self.calendar.new_events(f'{name[1]}', f'Время: {name[0]}: {name[2]}', f'{days[-1]}', f'{name[0]}', '7', groups)
                elif len(name) == 2 and name[-1] == 'Классный час':
                    print(f'{name[1]}', f'Время: {name[0]}', f'{days[-1]}', f'{name[0]}', '1', groups)
                    self.calendar.new_events(f'{name[1]}', f'Время: {name[0]}', f'{days[-1]}', f'{name[0]}', '1', groups)
                elif len(name) == 2:
                    print(f'{name[1]}', f'Время: {name[0]}', f'{days[-1]}', f'{name[0]}', '1', groups)
                    self.calendar.new_events(f'{name[1]}', f'Время: {name[0]}', f'{days[-1]}', f'{name[0]}', '1', groups)
            except Exception as e:
                self.__log_bot.message_bot(f'Возникла ошибка: {e}')
        self.__log_bot.message_bot('Программа заполнения расписания для групп закончила работу')

    def par_teacher(self):  # Парсинг расписания преподователей
        """Функция достает расписание преподователей с сайта, обрабатывает и
        вызывает функцию для загрузки их в календарь"""
        # self.__db.connection()
        self.__log_bot.message_bot('Парсинг расписания преподователей  и занесения в календарь начал свою работу')
        res = requests.get(self.url_site)
        soup = BeautifulSoup(res.text, 'lxml')
        lecturer = soup.find('select', id='teacher').find_all('option')
        clear_list = []
        preps = []
        for i in lecturer:
            clear_list.append(i.text)  # Засовываем все группы в список

        for teacher in self.name_teacher:
            teac = f'{teacher[:-5]} {teacher[-5:]}'
            if teacher in clear_list or teac in clear_list:
                trash_prep_soup = ''
                if teacher in clear_list:
                    trash_response_text = requests.post(self.url_site, {'teacher': f'{teacher}'.encode('cp1251'), 'stp': 'Показать!'.encode('cp1251')}).text
                    trash_prep_soup = BeautifulSoup(trash_response_text, 'lxml')
                elif teac in clear_list:
                    trash_response_text = requests.post(self.url_site, {'teacher': f'{teac}'.encode('cp1251'), 'stp': 'Показать!'.encode('cp1251')}).text
                    trash_prep_soup = BeautifulSoup(trash_response_text, 'lxml')

                teachers_list = []
                for i in trash_prep_soup.find_all(['p', 'b']):
                    teachers_list.append(str(i).replace('<br/><b>', '||').replace('</b><br/>', '||').replace('<p>', '').replace('<b>', '').replace('</p>', '').replace('</b>', ''))

                teacher = teacher.replace('  ', ' ')
                preps.append(f'&{teacher}')
                for i in teachers_list:
                    if '</' not in i:
                        preps.append(i.split('||'))


        today = str(date.today())  # Текущая дата
        days = list()  # Засовываем дату (год-месяц-день) текущей пары
        groups = str()  # Для хранения группы

        for name in preps:  # Обрабатываем данные
            if name[0] == "&":
                groups = f"{name[1:]}"
            if 12 <= len(name[0]) <= 23 and name[0][0:1].isdigit():
                nd = name[0].split(" ")  # Разделяем дату на 'день', 'месяц', 'день недели'
                nd[1] = nd[1].replace(",", "")  # Избавляемся от знака , в месяце и засовываем в список
                month = {'января': '01', 'февраля': '02', 'марта': '03', 'апреля': '04', 'мая': '05', 'июня': '06',
                         'июля': '07', 'августа': '08', 'сентября': '09', 'октября': '10', 'ноября': '11', 'декабря': '12'}
                if len(nd[0]) == 1 and 1 <= int(nd[0]) < 10:
                    days.append(f"{today[0:4]}-{month[nd[1]]}-0{nd[0]}")
                elif len(nd[0]) == 2 and int(nd[0]) >= 10:
                    days.append(f"{today[0:4]}-{month[nd[1]]}-{nd[0]}")

            try:
                if len(name) == 3:  # Обработанные данные загружаем в календарь
                    self.calendar.new_events(f'{name[1]}', f'Время: {name[0]}: {name[2]}', f'{days[-1]}', f'{name[0]}', '9', groups)
            except Exception as e:
                self.__log_bot.message_bot(f'Возникла ошибка: {e}')
        self.__log_bot.message_bot('Программа заполнения расписания для преподователей закончила работу')

    def planshet(self):
        """Функция достает и обрабатывает данные с планшетки и запускает функцию удаления ивентов"""
        filename = 'planchette.xlsx'
        wb1 = openpyxl.load_workbook(filename)
        couple = '1 пара'
        worksheet = wb1[couple]  # Открываем листы в планшетке
        dates = str(worksheet['B1'].value).split() if worksheet.max_column == 9 else str(worksheet['A1'].value).split()  # Достаем дату
        if '-' in dates[0]:
            dates = dates[0]
        else:
            dates = dates[0].split('.')
            dates = f'{dates[2]}-{dates[1]}-{dates[0]}'
        self.calendar.get_events_list(dates)

    def planchette(self):
        '''Функция достает и обрабатывает данные с планшетки и запускает функцию загрузки их в календарь'''
        #self.__log_bot.message_bot('Я начал парсить планшетку')
        wb1 = openpyxl.load_workbook('planchette.xlsx')
        days = datetime.today().isoweekday()
        spisok = list()
        if days == 1:
            couple_list = ['1 пара', '2 пара', '3 пара', '4 пара', '5 пара', '6 пара', 'Классные часы']
        else:
            couple_list = ['1 пара', '2 пара', '3 пара', '4 пара', '5 пара', '6 пара', '7 пара']
        try:
            for couple in couple_list:
                worksheet = wb1[couple]

                if worksheet.max_column == 9:
                    dates = str(worksheet['B1'].value).split()
                    if '-' in dates[0]:
                        datess = dates[0]
                        dates = dates[0].split('-')
                    else:
                        dates = dates[0].split('.')
                        datess = f'{dates[2]}-{dates[1]}-{dates[0]}'

                    couple_time = worksheet['A1'].value

                elif worksheet.max_column == 8:
                    dates = str(worksheet['A1'].value).split()
                    if '-' in dates[0]:
                        datess = dates[0]
                        dates = dates[0].split('-')
                        dates = [dates[-1], dates[1], dates[0]]
                    else:
                        dates = dates[0].split('.')
                        datess = f'{dates[2]}-{dates[1]}-{dates[0]}'

                    time1 = {'1 пара': '08:00-09:30', '2 пара': '09:40-11:10', '3 пара': '11:30-13:00',
                             '4 пара': '13:40-15:10', '5 пара': '15:30-17:00', '6 пара': '17:10-18:40',
                             'Классный час': '13:05-13:35', 'Классные часы': '13:05-13:35'}
                    time = {'1 пара': '08:00-09:30', '2 пара': '09:40-11:10', '3 пара': '11:30-13:00',
                            '4 пара': '13:10-14:40', '5 пара': '15:00-16:30', '6 пара': '16:40-18:10',
                            '7 пара': '18:20-19:50'}

                    try:
                        if days == 1:
                            couple_time = time1[couple]
                        else:
                            couple_time = time[couple]
                    except:
                        pass
                for j in range(0, worksheet.max_row):
                    my_list_1 = []
                    my_list_2 = []

                    if worksheet.max_column == 9:
                        for col in worksheet.iter_cols(2, 5):
                            a = col[j].value
                            if a != None:
                                my_list_1.append(a)

                    elif worksheet.max_column == 8:
                        for col in worksheet.iter_cols(1, 4):
                            a = col[j].value
                            if a != None:
                                my_list_1.append(a)

                    if len(my_list_1) == 3:
                        if '/' in my_list_1[1]:
                            for k in my_list_1[1].split('/'):
                                if 'КПК' not in my_list_1 and f'{dates[0]}.{dates[1]}.{dates[2]}' not in my_list_1:
                                    spisok.append([[couple_time], ['Без названия', f'{couple_time}: {my_list_2[2]} ауд {my_list_2[0]}', datess, couple_time, 7, my_list_2[1]]])
                        else:
                            if 'КПК' not in my_list_1 and f'{dates[0]}.{dates[1]}.{dates[2]}' not in my_list_1:
                                spisok.append([[couple_time], ['Без названия', f'{couple_time}: {my_list_1[2]} ауд {my_list_1[0]}', datess, couple_time, 7, my_list_1[1]]])

                    elif len(my_list_1) >= 4:
                        if '/' in my_list_1[1]:
                            for k in my_list_1[1].split('/'):
                                spisok.append([[couple_time], [my_list_1[3], f'{couple_time}: {my_list_1[2]} ауд {my_list_1[0]}', datess, couple_time, 9, k]])
                        else:
                            spisok.append([[couple_time], [my_list_1[3], f'{couple_time}: {my_list_1[2]} ауд {my_list_1[0]}', datess, couple_time, 9, my_list_1[1]]])
                    if worksheet.max_column == 9:
                        for col in worksheet.iter_cols(6, 9):
                            a = col[j].value
                            if a != None:
                                my_list_2.append(a)

                    elif worksheet.max_column == 8:
                        for col in worksheet.iter_cols(5, 8):
                            a = col[j].value

                            if a != None:
                                my_list_2.append(a)

                    if len(my_list_2) == 3:
                        if '/' in my_list_2[1]:
                            for j in my_list_1[2].split('/'):
                                if 'КПК' not in my_list_2 and f'{dates[0]}.{dates[1]}.{dates[2]}' not in my_list_2:
                                    spisok.append([[couple_time], ['Без названия', f'{couple_time}: {my_list_2[2]} ауд {my_list_2[0]}', datess, couple_time, 7, my_list_2[1]]])
                        else:
                            if 'КПК' not in my_list_2 and f'{dates[0]}.{dates[1]}.{dates[2]}' not in my_list_2:
                                spisok.append([[couple_time], ['Без названия', f'{couple_time}: {my_list_2[2]} ауд {my_list_2[0]}', datess, couple_time, 7, my_list_2[1]]])
                    elif len(my_list_2) >= 4:
                        if '/' in my_list_2[1]:
                            for j in my_list_2[1].split('/'):
                                spisok.append([[couple_time], [my_list_2[3], f'{couple_time}: {my_list_2[2]} ауд {my_list_2[0]}', datess, couple_time, 9, j]])
                        else:
                            spisok.append([[couple_time], [my_list_2[3], f'{couple_time}: {my_list_2[2]} ауд {my_list_2[0]}', datess, couple_time, 7, my_list_2[1]]])

            date = datetime.now()
            date = str(date.time()).split(':')
            Nowtime = f'{int(date[0])}:{date[1]}'
            self.calendar.delete_events()
            for j in spisok:
                time = j[0][0].split('-')
                if Nowtime <= time[1]:
                    groups = j[1][-1].strip()
                    if groups in self.name_groups:
                        text = j[1][1].split(': ')
                        text1 = text[-1].split('. ')
                        teacher = text1[-1]
                        teacher1 = f'{text1[0]}.'
                        kabinet = teacher.split()[-1].replace('.0', '')
                        self.events.append([kabinet, j[1][0], groups, teacher1, j[1][2], j[1][3]])
                        self.calendar.new_events(text=f'{j[1][0]}',
                                                 description=f'Время: {j[1][1]}',
                                                 day=f'{j[1][2]}',
                                                 time=f'{j[1][3]}',
                                                 color='7',
                                                 group=groups)
                    else:
                        self.__log_bot.message_bot(f'''Таких данных нет парсинг планшетки планшетки(2)
                                                     Данные которых нет в списке: \'{groups}\'''')
            for j in spisok:
                time = j[0][0].split('-')
                if Nowtime <= time[1]:
                    teacher = j[1][1].split('. ')
                    teacher = teacher[0].split(': ')
                    teacher = teacher[1]

                    try:
                        text = j[1][1].split(': ')
                        text1 = text[-1].split('. ')
                        teacher = teacher.strip()
                        teacher = teacher.replace('  ', ' ')
                        print(f"{j[1][0]} \n"
                              f"Время: {text[0]}: {j[1][-1]} {text1[-1]} \n"
                              f"{j[1][2]} \n"
                              f"{j[1][3]} \n"
                              f"9 \n"
                              f"{teacher}.")
                        # self.calendar.new_events(text=f'{j[1][0]}',
                        #                          description=f'Время: {text[0]}: {j[1][-1]} {text1[-1]}',
                        #                          day=f'{j[1][2]}',
                        #                          time=f'{j[1][3]}',
                        #                          color='9',
                        #                          group=f'{teacher}.')
                    except Exception as e:
                        self.__log_bot.message_bot(f'''Была вызвана ошибка в разделе парсинга планшетки(2)
                                                     Данные вызванные ошибку: {j}
                                                     описание ошибки: {e}''')
            self.__log_bot.message_bot('Расписание с планшетки было занесено в календарь')
        except Exception as e:
            self.__log_bot.message_bot(f'''Была вызвана ошибка в разделе парсинга планшетки(2.1)
                                     описание ошибки: {e}''')
        #self.new_sql_events()

    def new_sql_events(self):
        return None
        try:
            for i in self.events:
                date = self.__db.select_db_data(group=i[2], date=i[4], time=i[5])
                if date != None:
                    if i[2] == date['Group'] and i[4] == date['Couple_date'] and i[5] == date['Couple_time'] and i[0] != date['Cabinet'] or i[1] != date['Couple'] or i[3] != date['Teacher']:
                        self.__db.update_person_date(cabinet=i[0], couple=i[1], teacher=i[3], id=date['ID'])
                    else:
                        pass
                else:
                    self.__db.insert_db_person(Cabinet=i[0], Couple=i[1], Teacher=i[3], Couple_date=i[4], Couple_time=i[5], Group=i[2])
                    pass
        except Exception as e:
            self.__log_bot.message_bot(f'Произошла ошибка при занесении расписания в БД \nОшибка: {e}')
        finally:
            self.__db.close_connection()
            self.events.clear()

    def planchette_check(self) -> bool:
        """Функция проверяет соответствует ли планшетка требованиям"""
        workbook_name = 'planchette.xlsx'
        wb1 = openpyxl.load_workbook(workbook_name)
        worksheet = wb1['1 пара'].max_column

        if worksheet == 9:
            self.planshet()
            return True
        elif worksheet == 8:
            self.__log_bot.message_bot('У планшетки отсувствует время пары, ошибка будет отработана')
            self.planshet()
            return True
        else:
            self.__log_bot.message_bot('Планшетка не соответствует требованиям')
            return False
